from __future__ import annotations

import csv
import hashlib
import io
from pathlib import Path
from typing import Any

from ..markdown import normalize_markdown
from ..models import KnowledgeDocument
from .base import DocumentProcessor


def _render_markdown_table(rows: list[list[str]], max_rows: int = 1000, max_cols: int = 50) -> str:
    if not rows:
        return "*[Tabela vazia]*"

    # Filter empty trailing rows and columns
    cleaned_rows = []
    for row in rows[:max_rows]:
        cleaned_row = [str(cell if cell is not None else "").strip().replace("\n", " ").replace("|", "\\|") for cell in row[:max_cols]]
        if any(cleaned_row):
            cleaned_rows.append(cleaned_row)

    if not cleaned_rows:
        return "*[Tabela vazia]*"

    # Determine column count
    col_count = max(len(row) for row in cleaned_rows)
    if col_count == 0:
        return "*[Tabela vazia]*"

    header = cleaned_rows[0] + [""] * (col_count - len(cleaned_rows[0]))
    # If header contains only numbers/empty, generate Column 1, Column 2...
    if not any(header):
        header = [f"Coluna {i+1}" for i in range(col_count)]

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * col_count) + " |",
    ]

    data_rows = cleaned_rows[1:] if len(cleaned_rows) > 1 else []
    for row in data_rows:
        padded = row + [""] * (col_count - len(row))
        lines.append("| " + " | ".join(padded[:col_count]) + " |")

    if len(rows) > max_rows:
        lines.append(f"\n*[Exibindo primeiras {max_rows} linhas de {len(rows)}]*")

    return "\n".join(lines)


class SpreadsheetProcessor(DocumentProcessor):
    """Processor for spreadsheets (XLSX, XLS, XLSM without macros, CSV, TSV, ODS) to Markdown tables."""

    name = "spreadsheet"
    supported_extensions = (".xlsx", ".xls", ".xlsm", ".csv", ".tsv", ".ods")
    supported_mimetypes = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
        "text/tab-separated-values",
        "application/vnd.oasis.opendocument.spreadsheet",
    )

    def process_file(
        self,
        file_path: Path | str,
        *,
        metadata: dict[str, Any] | None = None,
        options: dict[str, Any] | None = None,
    ) -> KnowledgeDocument:
        path = Path(file_path)
        content_bytes = path.read_bytes()
        doc_metadata = dict(metadata or {})
        doc_metadata.setdefault("file_path", str(path.resolve()))
        doc_metadata.setdefault("filename", path.name)
        return self.process_bytes(
            content_bytes,
            filename=path.name,
            mime_type="",
            metadata=doc_metadata,
            options=options,
        )

    def process_bytes(
        self,
        content_bytes: bytes,
        *,
        filename: str = "",
        mime_type: str = "",
        metadata: dict[str, Any] | None = None,
        options: dict[str, Any] | None = None,
    ) -> KnowledgeDocument:
        del mime_type
        doc_metadata = dict(metadata or {})
        doc_options = dict(options or {})
        doc_id = hashlib.sha256(content_bytes).hexdigest()
        title = doc_metadata.get("title") or Path(filename).stem or "Planilha"
        ext = Path(filename).suffix.lower() if filename else ".xlsx"

        max_rows = int(doc_options.get("max_rows_per_sheet", 1000))
        max_cols = int(doc_options.get("max_cols_per_sheet", 50))

        sections: list[str] = []
        sheet_names: list[str] = []

        if ext in {".csv", ".tsv"}:
            delimiter = "\t" if ext == ".tsv" else ","
            try:
                text_content = content_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                text_content = content_bytes.decode("latin-1", errors="replace")

            reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
            rows = [list(row) for row in reader]
            sheet_table = _render_markdown_table(rows, max_rows=max_rows, max_cols=max_cols)
            sections.append(sheet_table)
            sheet_names.append("Dados")
        else:
            # Excel XLSX / XLSM (never execute macros, data_only=True)
            try:
                import openpyxl

                wb = openpyxl.load_workbook(
                    io.BytesIO(content_bytes),
                    data_only=True,
                    read_only=True,
                    keep_vba=False,
                )
                try:
                    sheet_names = wb.sheetnames
                    for sheet_name in sheet_names:
                        ws = wb[sheet_name]
                        sheet_rows: list[list[str]] = []
                        for row in ws.iter_rows(values_only=True):
                            sheet_rows.append(
                                [str(cell) if cell is not None else "" for cell in row]
                            )
                            if len(sheet_rows) >= max_rows + 10:
                                break

                        table_md = _render_markdown_table(
                            sheet_rows, max_rows=max_rows, max_cols=max_cols
                        )
                        sections.append(f"## Planilha: {sheet_name}\n\n{table_md}")
                finally:
                    wb.close()
            except Exception as exc:
                sections.append(f"*[Não foi possível processar planilha: {exc}]*")

        full_content = normalize_markdown("\n\n".join(sections))
        doc_metadata.setdefault("raw_type", "spreadsheet")
        doc_metadata.setdefault("sheet_names", sheet_names)
        doc_metadata.setdefault("filename", filename)

        return KnowledgeDocument(
            id=doc_id,
            container_id=str(doc_metadata.get("container_id") or "spreadsheets"),
            title=title,
            content=full_content,
            original_url=str(doc_metadata.get("original_url") or doc_metadata.get("file_path") or filename),
            source_type=str(doc_metadata.get("source_type") or "local_files"),
            container_name=str(doc_metadata.get("container_name") or "Planilhas"),
            path=list(doc_metadata.get("path") or [title]),
            metadata=doc_metadata,
        )
